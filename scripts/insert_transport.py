"""
Re-insert transport data from Excel into Supabase.

Usage:
    python scripts/insert_transport.py <path_to_excel>

Handles both named and unnamed column headers robustly.
After running the 0005 migration in Supabase, re-run with --full
to also populate supplier / from_location / to_location / services.
"""

import sys
import re
import json
import datetime
import warnings
import requests
import openpyxl

warnings.filterwarnings("ignore")

# ── Supabase config ──────────────────────────────────────────────────────────
SUPABASE_URL = "https://nqrzhtxtlagevmtprvlt.supabase.co"
SUPABASE_KEY = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9."
    "eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im5xcnpodHh0bGFnZXZtdHBydmx0Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NDk3OTg0MiwiZXhwIjoyMDkwNTU1ODQyfQ."
    "6lbh3tNY2M9y52Cu58w76adLnMzJTIOetZ9mYMMwA6I"
)
BASE_HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
}

# ── Terminal map ─────────────────────────────────────────────────────────────
TERMINAL_MAP = [
    (re.compile(r"singapore airlines|silkair|\bsq\b", re.I), "T2/T3"),
    (re.compile(r"scoot|\btr\b", re.I), "T1"),
    (re.compile(r"jetstar|\b3k\b|\bjq\b", re.I), "T4"),
    (re.compile(r"cathay|\bcx\b", re.I), "T4"),
    (re.compile(r"emirates|\bek\b", re.I), "T1"),
    (re.compile(r"qatar|\bqr\b", re.I), "T1"),
    (re.compile(r"malaysia|\bmy\b|\bmh\b", re.I), "T1"),
    (re.compile(r"thai|\btg\b", re.I), "T1"),
    (re.compile(r"air asia|\bak\b|\bfd\b", re.I), "T4"),
    (re.compile(r"batik|\bod\b", re.I), "T4"),
    (re.compile(r"lion air|\bjt\b", re.I), "T4"),
    (re.compile(r"garuda|\bga\b", re.I), "T3"),
    (re.compile(r"korean|\bke\b", re.I), "T4"),
    (re.compile(r"japan airlines|\bjl\b", re.I), "T1"),
    (re.compile(r"ana|\bnh\b", re.I), "T1"),
    (re.compile(r"lufthansa|\blh\b", re.I), "T1"),
    (re.compile(r"british|\bba\b", re.I), "T1"),
    (re.compile(r"china southern|\bcz\b", re.I), "T2"),
    (re.compile(r"china eastern|\bmu\b", re.I), "T2"),
    (re.compile(r"air china|\bca\b", re.I), "T2"),
    (re.compile(r"\baf\b|air france", re.I), "T1"),
    (re.compile(r"\bua\b|united", re.I), "T1"),
    (re.compile(r"\bkl\b|klm", re.I), "T1"),
    (re.compile(r"\btk\b|turkish", re.I), "T1"),
    (re.compile(r"\bku\b|kuwait", re.I), "T1"),
]


def get_terminal(s: str) -> str:
    if not s:
        return "TBC"
    for pat, term in TERMINAL_MAP:
        if pat.search(s):
            return term
    return "TBC"


def infer_type(services: str, from_: str, to: str) -> str:
    svc = (services or "").lower()
    frm = (from_ or "").lower()
    to_ = (to or "").lower()
    if re.search(r"\barrival\b|\barr\b", svc):
        return "Arrival"
    if re.search(r"\bdeparture\b|\bdep\b", svc):
        return "Departure"
    from_is_airport = bool(re.search(r"changi|airport|terminal", frm))
    to_is_airport   = bool(re.search(r"changi|airport|terminal", to_))
    if from_is_airport and not to_is_airport:
        return "Arrival"
    if to_is_airport and not from_is_airport:
        return "Departure"
    if re.search(r"airport.*to.*hotel|changi.*to.*hotel", svc):
        return "Arrival"
    if re.search(r"hotel.*to.*airport|hotel.*to.*changi", svc):
        return "Departure"
    return "Arrival"


SGT = datetime.timezone(datetime.timedelta(hours=8))


def to_sgt_iso(date_val, time_val) -> str:
    if isinstance(date_val, datetime.datetime):
        d = date_val.date()
    elif isinstance(date_val, datetime.date):
        d = date_val
    else:
        return datetime.datetime.now(tz=datetime.timezone.utc).isoformat()

    h, m = 0, 0
    if isinstance(time_val, datetime.time):
        h, m = time_val.hour, time_val.minute
    elif isinstance(time_val, datetime.datetime):
        h, m = time_val.hour, time_val.minute
    elif isinstance(time_val, str):
        match = re.match(r"(\d{1,2}):(\d{2})", time_val.strip())
        if match:
            h, m = int(match.group(1)), int(match.group(2))

    sgt_dt = datetime.datetime(d.year, d.month, d.day, h, m, 0, tzinfo=SGT)
    return sgt_dt.astimezone(datetime.timezone.utc).isoformat().replace("+00:00", "Z")


def excel_date_str(date_val) -> str:
    if isinstance(date_val, datetime.datetime):
        return date_val.date().isoformat()
    if isinstance(date_val, datetime.date):
        return date_val.isoformat()
    return ""


def safe_str(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    # Replace common encoding garbage chars
    s = s.replace("\ufffd", "").strip()
    return s


def safe_int(v) -> int:
    try:
        return max(1, int(v))
    except (TypeError, ValueError):
        return 1


def normalise_driver(v: str) -> str | None:
    """Return None for blank/N/A, otherwise return cleaned string."""
    s = safe_str(v).strip()
    if not s or re.fullmatch(r"n/?a", s, re.I):
        return None
    return s


# ── Column picker ─────────────────────────────────────────────────────────────
# Keys are lowercase to match normalised headers.
COL = {
    "date":         ["date"],
    "agent":        ["agent", "airline"],
    "file_ref":     ["file ref", "fileref", "ref", "file_ref", "id"],
    "pax_name":     ["passenger name", "fila name", "pax name", "passenger", "client", "name"],
    "pax_count":    ["total pax", "pax", "pax count", "passengers", "no. of pax"],
    "pickup_time":  ["p.up/eta", "pickup", "eta", "p.up"],
    "dropoff_time": ["d.off/etd", "dropoff", "drop off", "etd", "d.off"],
    "flight":       ["flight details", "flight", "flight no"],
    "driver":       ["driver contact", "driver name & contact", "driver", "driver name"],
    "terminal":     ["terminal", "t"],   # "t" covers single-letter "T" header
    "services":     ["services", "service"],
    "from":         ["from"],
    "to":           ["to"],
    "supplier":     ["supplier"],
}

# Known positional roles for columns with empty headers (0-indexed)
# If a column header is blank we fall back to this map.
POSITIONAL_ROLE = {
    10: "driver",   # Col K — driver contact (used in Apr 8-14 template)
}


def pick(row_lower: dict, row_positional: list, field: str):
    """Look up a field value, trying named aliases first then positional fallback."""
    for alias in COL[field]:
        v = row_lower.get(alias)
        if v is not None and v != "":
            return v
    # Positional fallback: scan POSITIONAL_ROLE for this field
    for idx, role in POSITIONAL_ROLE.items():
        if role == field and idx < len(row_positional):
            v = row_positional[idx]
            if v is not None and v != "":
                return v
    return None


def parse_sheet(path: str, full_mode: bool = False):
    wb = openpyxl.load_workbook(path, data_only=True)
    names_lower = {n.lower(): n for n in wb.sheetnames}
    sheet_name = names_lower.get("transport", wb.sheetnames[0])
    ws = wb[sheet_name]
    print(f"Using sheet: '{sheet_name}'  ({ws.max_row - 1} data rows)")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("Sheet is empty.")
        return []

    raw_headers = rows[0]
    # Normalise headers to lowercase; empty cells become empty string ""
    headers_norm = [str(h).strip().lower() if h is not None else "" for h in raw_headers]

    records = []
    for row_idx, row_vals in enumerate(rows[1:], start=2):
        row_lower = {}
        for i, v in enumerate(row_vals):
            h = headers_norm[i] if i < len(headers_norm) else ""
            if h:  # only store non-empty header keys
                row_lower[h] = v
        row_pos = list(row_vals)  # positional fallback

        file_ref = safe_str(pick(row_lower, row_pos, "file_ref"))
        if not file_ref or file_ref.lower() in ("", "none", "nan"):
            continue

        date_val    = pick(row_lower, row_pos, "date")
        pickup_val  = pick(row_lower, row_pos, "pickup_time")
        dropoff_val = pick(row_lower, row_pos, "dropoff_time")
        services    = safe_str(pick(row_lower, row_pos, "services"))
        from_       = safe_str(pick(row_lower, row_pos, "from"))
        to          = safe_str(pick(row_lower, row_pos, "to"))
        flight      = safe_str(pick(row_lower, row_pos, "flight"))
        agent       = safe_str(pick(row_lower, row_pos, "agent"))
        terminal_r  = safe_str(pick(row_lower, row_pos, "terminal"))
        pax_name    = safe_str(pick(row_lower, row_pos, "pax_name")) or "Unknown Pax"
        pax_count   = safe_int(pick(row_lower, row_pos, "pax_count"))
        driver_raw  = safe_str(pick(row_lower, row_pos, "driver"))
        supplier    = safe_str(pick(row_lower, row_pos, "supplier"))

        transfer_type = infer_type(services, from_, to)
        time_val      = pickup_val if transfer_type == "Arrival" else (dropoff_val or pickup_val)
        scheduled_time = to_sgt_iso(date_val, time_val)
        date_str = excel_date_str(date_val)
        if not date_str and scheduled_time:
            utc_dt = datetime.datetime.fromisoformat(scheduled_time.replace("Z", "+00:00"))
            date_str = utc_dt.astimezone(SGT).date().isoformat()

        terminal   = terminal_r or get_terminal(agent or flight)
        driver_info = normalise_driver(driver_raw)

        rec = {
            "file_ref":       file_ref,
            "date":           date_str,
            "pax_name":       pax_name,
            "pax_count":      pax_count,
            "flight_number":  flight or None,
            "agent":          agent or None,
            "terminal":       terminal,
            "type":           transfer_type,
            "scheduled_time": scheduled_time,
            "updated_time":   None,
            "driver_info":    driver_info,
            "notified":       False,
        }
        if full_mode:
            rec["services"]      = services or None
            rec["from_location"] = from_ or None
            rec["to_location"]   = to or None
            rec["supplier"]      = supplier or None

        records.append(rec)
        driver_display = driver_info or "—"
        terminal_display = terminal or "TBC"
        print(
            f"  Row {row_idx:3d}: {file_ref:<14} {transfer_type:9} "
            f"T={terminal_display:<7} driver={driver_display:<30} {pax_name[:35]}"
        )

    return records


def deduplicate(records):
    seen = {}
    for r in records:
        key = f"{r['file_ref']}|{r['scheduled_time']}"
        seen[key] = r
    return list(seen.values())


def snapshot_notified(file_refs):
    r = requests.get(
        f"{SUPABASE_URL}/rest/v1/flights",
        headers=BASE_HEADERS,
        params={
            "select": "file_ref,scheduled_time",
            "file_ref": f"in.({','.join(file_refs)})",
            "notified": "eq.true",
        },
    )
    if r.status_code != 200:
        print(f"Warning: could not snapshot notified rows: {r.text}")
        return set()
    return {f"{row['file_ref']}|{row['scheduled_time']}" for row in r.json()}


def delete_existing(file_refs):
    r = requests.delete(
        f"{SUPABASE_URL}/rest/v1/flights",
        headers=BASE_HEADERS,
        params={"file_ref": f"in.({','.join(file_refs)})"},
    )
    if r.status_code not in (200, 204):
        raise RuntimeError(f"Delete failed ({r.status_code}): {r.text}")
    print(f"  Deleted existing rows for {len(file_refs)} file refs.")


def insert_records(records):
    r = requests.post(
        f"{SUPABASE_URL}/rest/v1/flights",
        headers={**BASE_HEADERS, "Prefer": "return=minimal"},
        json=records,
    )
    if r.status_code not in (200, 201, 204):
        raise RuntimeError(f"Insert failed ({r.status_code}): {r.text}")
    print(f"  Inserted {len(records)} records.")


def main():
    args = sys.argv[1:]
    full_mode = "--full" in args
    paths = [a for a in args if not a.startswith("--")]

    if not paths:
        paths = [
            r"c:\Users\aariz\Downloads\1_-7_APR_RC_updated.xlsx",
            r"c:\Users\aariz\Downloads\OPS -8_-14_APR_RC.xlsx",
        ]

    mode_label = "FULL (with supplier/from/to/services)" if full_mode else "BASE (driver + all core fields)"
    print(f"Mode: {mode_label}\n")

    # ── Collect ALL records from ALL files first ─────────────────────────────
    # This is critical: if the same file_ref appears in multiple weekly files
    # (e.g. a booking that spans two weeks), we must process them together.
    # Processing files one-by-one with delete-then-insert would lose records
    # from the first file when the second file's delete runs.
    all_records: list[dict] = []

    for path in paths:
        print(f"\n{'='*60}")
        print(f"Parsing: {path}")
        records = parse_sheet(path, full_mode=full_mode)
        if records:
            all_records.extend(records)
            print(f"  Parsed {len(records)} rows from this file.")
        else:
            print("  No valid records.")

    if not all_records:
        print("\nNo records found across any file. Aborting.")
        return

    # ── Single batch: deduplicate across ALL files ───────────────────────────
    # When the same (file_ref, scheduled_time) appears in both files, keep the
    # LATER file's version (it's more up to date).
    deduped = deduplicate(all_records)
    print(f"\n{'='*60}")
    print(f"Total records across all files (after dedup): {len(deduped)}")

    # ── One delete + one insert for everything ───────────────────────────────
    file_refs = list({r["file_ref"] for r in deduped})
    print(f"Unique file refs: {len(file_refs)}")

    notified_set = snapshot_notified(file_refs)
    print(f"Previously notified pairs: {len(notified_set)}")

    for r in deduped:
        if f"{r['file_ref']}|{r['scheduled_time']}" in notified_set:
            r["notified"] = True

    delete_existing(file_refs)
    insert_records(deduped)
    print("\nDone!")


if __name__ == "__main__":
    main()
